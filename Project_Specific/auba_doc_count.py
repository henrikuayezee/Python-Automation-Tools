"""
Script: auba_doc_count.py
Description: Tool for auba doc count
Category: Project_Specific
"""
import os
import csv
from PyPDF2 import PdfReader
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm

def count_pdf_pages(file_path):
    try:
        with open(file_path, 'rb') as file:
            pdf_reader = PdfReader(file)
            return len(pdf_reader.pages)
    except Exception as e:
        print(f"Error reading PDF file {file_path}: {e}")
        return None

def count_excel_sheets(file_path):
    try:
        if file_path.endswith('.xlsx'):
            workbook = load_workbook(filename=file_path, read_only=True)
            return len(workbook.sheetnames)
        elif file_path.endswith('.xls'):
            workbook = pd.ExcelFile(file_path)
            return len(workbook.sheet_names)
        else:
            return None
    except Exception as e:
        print(f"Error reading Excel file {file_path}: {e}")
        return None

def main(directory_path, output_csv):
    files_data = []
    all_files = [f for f in os.listdir(directory_path) if f.endswith('.pdf') or f.endswith(('.xlsx', '.xls'))]
    
    with tqdm(total=len(all_files), desc="Processing files") as pbar:
        for file_name in all_files:
            file_path = os.path.join(directory_path, file_name)
            if file_name.endswith('.pdf'):
                pages = count_pdf_pages(file_path)
                files_data.append([file_name, 'PDF', pages])
            elif file_name.endswith(('.xlsx', '.xls')):
                sheets = count_excel_sheets(file_path)
                files_data.append([file_name, 'Excel', sheets])
            pbar.update(1)

    # Save data to CSV
    with open(output_csv, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(["File Name", "File Type", "Count (Pages/Sheets)"])
        writer.writerows(files_data)

if __name__ == "__main__":
    # Hardcoded directory path
    directory = r"C:\Users\henry\Downloads\Compressed\files-20241113T144803Z-001\files"  # Replace with your actual directory path
    output_csv_file = "output.csv"  # You can also change this to your desired output file name
    main(directory, output_csv_file)
