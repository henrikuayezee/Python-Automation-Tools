"""
Script: xlsx.py
Description: Tool for xlsx
Category: Data_Processing_and_Web
"""
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Set the directory path containing XLSX files
xlsx_directory = r"C:\Users\henry\Downloads\Compressed\Labeling-20230703T112211Z-001\Labeling"

# Set the output file path and name
output_file = r"C:\Users\henry\Downloads\Compressed\Labeling-20230703T112211Z-001\Labeling\output.xlsx"

# Initialize an empty list to hold all the data
all_data = []

# Loop through all the files in the directory
for filename in os.listdir(xlsx_directory):
    if filename.endswith('.xlsx'):
        # Open each XLSX file
        file_path = os.path.join(xlsx_directory, filename)
        workbook = load_workbook(file_path)
        # Read the data from each sheet in the XLSX file
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                all_data.append(row)

# Write the data to the output file
workbook = load_workbook()
worksheet = workbook.active
for row in all_data:
    worksheet.append(row)

workbook.save(output_file)
